# -*- coding=utf-8 -*-
import os
import tkinter.filedialog
import re
import time
import openpyxl
from openpyxl.styles import PatternFill  # , Border, Side, Alignment, Protection, Font,fills,colors


# from PyQt5.QtWidgets import QFileDialog

def getKeyWordFromSettingFile():
    global TER_keyword
    global ADV_keyword

    SettingFilePath = os.getcwd() + "\settings\key_word_definition.txt"
    print(SettingFilePath)
    # format:  Platform$__SiteNum$__PatName$__PlotStart
    TER_keyword = {'Platform': '', 'Item': '', 'SiteNum': '', 'PatName': '', 'PlotStart': '', 'PlotEnd': '',
                   'RowOffset': '', 'PassSymbol': '', 'FailSymbol': ''}
    ADV_keyword = {'Platform': '', 'Item': '', 'SiteNum': '', 'PatName': '', 'PlotStart': '', 'PlotEnd': '',
                   'RowOffset': '', 'PassSymbol': '', 'FailSymbol': ''}
    keyword = []

    file = open(SettingFilePath, 'r', encoding='utf-8')
    for line in file.readlines():
        if line[0] != '#' and line != '':
            print(line)
            keyword = line.split('$__')
            # print(keyword)
            if keyword[0] == 'TER':
                TER_keyword['Platform'] = keyword[0]
                TER_keyword['Item'] = keyword[1]
                TER_keyword['SiteNum'] = keyword[2]
                TER_keyword['PatName'] = keyword[3]
                TER_keyword['PlotStart'] = keyword[4]
                TER_keyword['PlotEnd'] = keyword[5]
                TER_keyword['RowOffset'] = keyword[6]
                TER_keyword['PassSymbol'] = 'P'  # keyword[7]
                TER_keyword['FailSymbol'] = '.'  # keyword[8]
                if len(keyword) == 9:
                    TER_keyword['PassSymbol'] = keyword[7]
                    TER_keyword['FailSymbol'] = keyword[8]
                # print(TER_keyword)
            if keyword[0] == 'ADV':
                ADV_keyword['Platform'] = keyword[0]
                ADV_keyword['Item'] = keyword[1]
                ADV_keyword['SiteNum'] = keyword[2]
                ADV_keyword['PatName'] = keyword[3]
                ADV_keyword['PlotStart'] = keyword[4]
                ADV_keyword['PlotEnd'] = keyword[5]
                ADV_keyword['RowOffset'] = keyword[6]
                ADV_keyword['PassSymbol'] = 'P'  # keyword[7]
                ADV_keyword['FailSymbol'] = '.'  # keyword[8]
                if len(keyword) == 9:
                    ADV_keyword['PassSymbol'] = keyword[7]
                    ADV_keyword['FailSymbol'] = keyword[8]
                # print(ADV_keyword)


def getDatalogInfo(TER_flag, ADV_flag):
    log_dir = os.getcwd()
    TER_files = ""
    ADV_files = ""
    sites_TER = []
    sites_ADV = []
    totalsiteCnt = 0
    str_sites_TER = []
    str_sites_ADV = []

    if TER_flag == 'TER':
        select_sites = ''

        TER_files = tkinter.filedialog.askopenfilenames(title=u"TER_选择文件",
                                                        initialdir=(os.path.expanduser(log_dir)))
        for each_file in TER_files:
            select_sites = str(input(each_file + '  ' + 'select sites:  '))
            if select_sites == "":
                for each_file in TER_files:
                    select_sites = (getAllSiteNums(each_file))
            totalsiteCnt = totalsiteCnt + len((select_sites.split(',')))
            str_sites_TER.append(select_sites)
    if ADV_flag == 'ADV':
        select_sites = ''
        ADV_files = tkinter.filedialog.askopenfilenames(title=u"ADV_选择文件",
                                                        initialdir=(os.path.expanduser(log_dir)))
        for each_file in ADV_files:
            print(each_file)
            select_sites = str(input(each_file + '  ' + 'select sites: '))

            if select_sites == "":
                for each_file in ADV_files:
                    select_sites = (getAllSiteNums(each_file))
            totalsiteCnt = totalsiteCnt + len((select_sites.split(',')))
            str_sites_ADV.append(select_sites)
    print("totalsiteCnt= " + str(totalsiteCnt))
    xls = openpyxl.Workbook()
    siteCnt = 0
    index = 0
    for each_file in TER_files:
        sites_TER = str_sites_TER[index].split(',')
        index = index + 1
        for each_site in sites_TER:
            siteCnt = siteCnt + 1
            processLog(each_file, each_site, xls, siteCnt, totalsiteCnt, TER_keyword)

    index = 0
    for each_file in ADV_files:
        for each_site in str_sites_ADV[index].split(','):
            sites_ADV = str_sites_ADV[index].split(',')
            siteCnt = siteCnt + 1
            processLog(each_file, each_site, xls, siteCnt, totalsiteCnt, ADV_keyword)

    time_flag = (time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time())))
    xls.save(os.getcwd() + '/shmplot_' + time_flag + '.xlsx')


def processLog(each_file, each_site, xls, siteCnt, totalsiteCnt, dict_keyword):
    # time_flag = (time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(time.time())))
    flag = 0
    startPlot = 0
    site_flag = 0
    each_item_info = []
    each_plot = []
    file = open(each_file, 'r', encoding='utf-8')
    shtName = 'Sheet'

    iRow = 2
    interval_columns = 20
    iColumn = (siteCnt - 1) * interval_columns
    for line in file.readlines():
        if ('VBT error' in line):
            continue
        if (dict_keyword['Item'] in line):
            each_item_info.append(line)
            flag = 1
        if (dict_keyword['SiteNum'] in line and not (
                (dict_keyword['SiteNum'] + str(each_site)) == line[:-1]) and flag == 1):
            each_item_info = []
            site_flag = 0

        if dict_keyword['Platform'] == 'TER':
            if (dict_keyword['SiteNum'] + str(each_site)) == line[:-1] and flag == 1:
                each_item_info.append(line)
                site_flag = 1
            if dict_keyword['PatName'] in line and flag == 1 and site_flag == 1:
                each_item_info.append(line)
        elif dict_keyword['Platform'] == 'ADV':
            if (dict_keyword['SiteNum'] + str(each_site)) == line[:-1] and flag == 1:
                each_item_info.append(line)
                site_flag = 1
            if dict_keyword['PatName'] in line and flag == 1 and site_flag == 1:
                each_item_info.append(line)
        if (line.strip().startswith(dict_keyword['PlotStart']) and flag == 1 and site_flag == 1):
            # (dict_keyword['PlotStart'] in line.strip()  and flag==1 and site_flag==1) :
            startPlot = 1
        if startPlot == 1 and flag == 1 and site_flag == 1:
            if line != '\n':
                each_plot.append(line)
        if (dict_keyword['PlotEnd'] in line and flag == 1 and startPlot == 1 and site_flag == 1):
            xls[shtName].cell(row=1, column=(siteCnt - 1) * interval_columns + 1, value=each_file)
            for tmpstr in each_item_info:
                iColumn = (siteCnt - 1) * interval_columns
                xls[shtName].cell(row=iRow + 1, column=iColumn + 1, value=tmpstr)
                xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + 1, value=tmpstr)
                iRow = iRow + 1
                pass
            for each_plot_line in each_plot:
                iColumn = (siteCnt - 1) * interval_columns
                tmpstr = each_plot_line.split('\t')
                tmpstr = each_plot_line.split()
                if dict_keyword['Platform'] == 'ADV' and not (ADV_keyword['PlotEnd'] in each_plot_line):
                    pass  # ADV_keyword['FailSymbol']
                    for i in range(1, len(tmpstr)):
                        tmp = tmpstr[i]
                        text_pass = ADV_keyword['PassSymbol']
                        text_fail = ADV_keyword['FailSymbol']
                        if re.match(text_pass, tmp) != None:
                            tmpstr[i] = 'P'
                        if re.match(text_fail, tmp) != None:
                            tmpstr[i] = '.'

                for x in tmpstr:
                    xls[shtName].cell(row=iRow + 1, column=iColumn + 1, value=x)
                    if x == 'P':
                        fill = PatternFill(start_color='00EE00', end_color='00EE00', fill_type="solid", )
                        xls[shtName].cell(row=iRow + 1, column=iColumn + 1).fill = fill
                    elif x == '.':
                        fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type="solid", )
                        xls[shtName].cell(row=iRow + 1, column=iColumn + 1).fill = fill
                    if str(xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                            iColumn - (siteCnt - 1) * interval_columns) + 1).value) != 'None':
                        y = str(xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                                    iColumn - (siteCnt - 1) * interval_columns) + 1).value) + x
                        xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                                    iColumn - (siteCnt - 1) * interval_columns) + 1, value=y)
                        if 'P' in y and '.' in y:
                            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid", )
                            xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                                        iColumn - (siteCnt - 1) * interval_columns) + 1).fill = fill
                    else:
                        xls[shtName].cell(row=iRow + 1, column=(totalsiteCnt * interval_columns) + (
                                    iColumn - (siteCnt - 1) * interval_columns) + 1, value=x)
                    iColumn = iColumn + 1
                iRow = iRow + 1
                pass

            startPlot = 0
            flag = 0
            site_flag = 0
            each_item_info = []
            each_plot = []
            if dict_keyword['Platform'] == 'TER':
                iRow = iRow + int(TER_keyword['RowOffset'])
            elif dict_keyword['Platform'] == 'ADV':
                iRow = iRow + int(ADV_keyword['RowOffset'])

    # xls.save(os.getcwd() + '/shmplot_'+time_flag+'.xlsx')


def getAllSiteNums(each_file):
    global TER_keyword
    global ADV_keyword
    site_info = []
    str_site = ''
    unique_site_info = []
    file = open(each_file, 'r', encoding='utf-8')
    for each_line in file.readlines():
        if TER_keyword['SiteNum'] in each_line:
            site_info.append(int(each_line[len(TER_keyword['SiteNum']):len(each_line)].strip()))
        elif ADV_keyword['SiteNum'] != "" and ADV_keyword['SiteNum'] in each_line:
            site_info.append(int(each_line[len(TER_keyword['SiteNum']):len(each_line)].strip()))
    unique_site_info = list(set(site_info))
    # unique_site_info=list.sort(unique_site_info)

    for x in unique_site_info:
        if str_site == '':
            str_site = str(x)
        else:
            str_site = str_site + ',' + str(x)

    return str_site


if __name__ == '__main__':
    global TER_keyword
    global ADV_keyword

    getKeyWordFromSettingFile()
    TER_flag = TER_keyword['Platform']
    ADV_flag = ADV_keyword['Platform']
    getDatalogInfo(TER_flag, ADV_flag)
